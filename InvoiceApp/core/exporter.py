"""导出模块 — 支持 Excel 和 CSV"""

import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .parser import BASIC_FIELDS, ITEM_FIELDS

# 样式
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_FONT = Font(name='微软雅黑', size=10)
CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center')
CELL_ALIGNMENT_CENTER = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


# ==================== 内部工具函数 ====================


def _write_header(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _format_sheet(ws, headers):
    """设置列宽、冻结、筛选"""
    col_count = len(headers)
    widths = {
        0: 30,
        1: 18,
        2: 16,
        3: 24,
        4: 16,
        5: 14,
        6: 24,
        7: 30,
        8: 24,
        9: 30,
        10: 30,
        11: 30,
        12: 24,
        13: 30,
        14: 30,
        15: 20,
        16: 12,
        17: 30,
        18: 12,
        19: 12,
        20: 12,
        21: 20,
        22: 12,
        23: 8,
        24: 8,
        25: 10,
        26: 10,
        27: 8,
        28: 10,
    }
    for i in range(col_count):
        col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
        ws.column_dimensions[col_letter].width = widths.get(i, 15)

    ws.freeze_panes = 'A2'

    last_col = chr(65 + col_count - 1) if col_count <= 26 else 'Z'
    ws.auto_filter.ref = f'A1:{last_col}{ws.max_row}'


def export_to_excel_grouped(results, output_path):
    """导出到Excel，项目明细展开为多行，发票级字段合并单元格

    results: [(data_dict, pdf_path), ...]  # 每张发票一个 dict（含 _items）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '发票信息'

    # 基本字段 + 项目字段 → 列定义
    top_fields = BASIC_FIELDS  # 发票级（合并单元格）
    item_fields = ITEM_FIELDS  # 项目级（逐行填写）
    all_fields = top_fields + item_fields
    headers = [f[0] for f in all_fields] + ['文件路径']
    top_count = len(top_fields)  # 前 N 列是发票级

    _write_header(ws, headers)

    current_row = 2  # 从第2行开始写数据
    for data_dict, pdf_path in results:
        if 'error' in data_dict:
            continue

        items = data_dict.get('_items', [])
        if not items:
            items = [{}]
        n = len(items)

        # 为这条发票预留 n 行
        for offset in range(n):
            row_idx = current_row + offset

            # 发票级字段只写第一行，其余行合并
            if offset == 0:
                for col, (_, key) in enumerate(top_fields, 1):
                    value = data_dict.get(key, '')
                    if isinstance(value, float):
                        value = round(value, 2)
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.font = CELL_FONT
                    cell.alignment = CELL_ALIGNMENT
                    cell.border = THIN_BORDER
            else:
                # 合并单元格
                for col in range(1, top_count + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = THIN_BORDER

            # 项目字段逐行填写
            item = items[offset]
            for col, (_, key) in enumerate(item_fields, top_count + 1):
                value = item.get(key, '')
                if isinstance(value, float):
                    value = round(value, 2)
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

            # 最后一列：文件路径
            path_col = len(all_fields) + 1
            cell = ws.cell(row=row_idx, column=path_col, value=str(pdf_path))
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        # 合并发票级单元格（如果有多个项目）
        if n > 1:
            for col in range(1, top_count + 1):
                ws.merge_cells(
                    start_row=current_row,
                    start_column=col,
                    end_row=current_row + n - 1,
                    end_column=col,
                )

        current_row += n

    _format_sheet(ws, headers)
    wb.save(output_path)
    return output_path


def export_to_csv(results, output_path, fields=None):
    """导出到 CSV，每张发票一行，项目明细合并为一个字段

    results: [(data_dict, pdf_path), ...]
    fields: 可选，要导出的字段列表 [(显示名, 键名), ...]，默认全部
    """
    from .parser import MERGE_FIELDS, merge_results

    all_fields = fields or MERGE_FIELDS
    headers = [f[0] for f in all_fields] + ['文件路径']
    merged = merge_results([(d, p) for d, p in results if 'error' not in d])

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for data_dict, pdf_path in merged:
            row = []
            for _, key in all_fields:
                value = data_dict.get(key, '')
                if isinstance(value, float):
                    value = round(value, 2)
                row.append(value)
            row.append(str(pdf_path))
            writer.writerow(row)

    return output_path
