import os
import re
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ========== 发票解析模块 ==========


def parse_invoice_text(text):
    """解析数电发票文本，返回字段字典"""
    data = {}
    lines = text.split('\n')

    # 发票号码
    m = re.search(r'发票号码[：:]\s*(\S+)', text)
    if m:
        data['发票号码'] = m.group(1).strip()

    # 开票日期
    m = re.search(r'开票日期[：:]\s*(\S+)', text)
    if m:
        data['开票日期'] = m.group(1).strip()

    # 购买方名称 & 销售方名称
    # 同一行格式: 买 名称：XXX 售 名称：YYY
    for line in lines:
        if '名称' in line and '售' in line:
            m_buyer = re.search(r'名称[：:]\s*(.+?)\s+售\s+名称[：:]', line)
            if m_buyer:
                data['购买方名称'] = m_buyer.group(1).strip()
            m_seller = re.search(r'售\s+名称[：:]\s*(.+)', line)
            if m_seller:
                data['销售方名称'] = m_seller.group(1).strip()
            break
    else:
        # 备用: 单独匹配
        m = re.search(r'名称[：:]\s*(.+?)\s*(?:统一社会信用|$)', text)
        if m:
            data['购买方名称'] = m.group(1).strip()
        m = re.search(r'名称[：:]\s*(.+?)$', text, re.MULTILINE)
        if m and '购买方名称' not in data:
            data['购买方名称'] = m.group(1).strip()

    # 纳税人识别号 - 按行提取，第一个是购买方，第二个是销售方
    tax_ids = re.findall(r'纳税人识别号[）\)]?[：:]\s*(\w+)', text)
    if len(tax_ids) >= 2:
        data['购买方纳税人识别号'] = tax_ids[0]
        data['销售方纳税人识别号'] = tax_ids[1]
    elif len(tax_ids) == 1:
        data['购买方纳税人识别号'] = tax_ids[0]

    # 项目名称
    m = re.search(r'\*经营租赁\*(\S+)', text)
    if m:
        data['项目名称'] = '*经营租赁*' + m.group(1).strip()

    # 金额（从项目明细行提取）
    for line in lines:
        if '*经营租赁*' in line:
            parts = line.split()
            if len(parts) >= 6:
                data['数量'] = parts[-5]
                data['单价'] = parts[-4]
                data['金额'] = parts[-3]
            break

    # 价税合计（小写）
    m = re.search(r'价税合计.*?小写[）\)]?\s*[¥￥]?\s*([\d.]+)', text)
    if m:
        data['价税合计'] = float(m.group(1))
    else:
        m = re.search(r'[¥￥]\s*([\d.]+)\s*(\*\*\*|\s*)$', text, re.MULTILINE)
        if m:
            data['价税合计'] = float(m.group(1))

    # 价税合计（大写）
    m = re.search(r'价税合计（大写）\s*(.+?)（小写）', text)
    if m:
        data['价税合计大写'] = m.group(1).strip()

    # 车牌号
    m = re.search(r'车牌号[：:]\s*(\S+)', text)
    if m:
        data['车牌号'] = m.group(1).strip()

    # 车辆类型
    m = re.search(r'车辆类型[：:]\s*(\S+)', text)
    if m:
        data['车辆类型'] = m.group(1).strip()

    # 通行日期起/止 - 取到行尾，包含中间空格
    for line in lines:
        if '通行日期起/止' in line:
            m = re.search(r'通行日期起/止[：:]\s*(.+)', line)
            if m:
                data['通行日期起止'] = m.group(1).strip()
            break

    # 入/出口站
    for line in lines:
        if '入/出口站' in line:
            m = re.search(r'入/出口站[：:]\s*(.+?)(?:（|$)', line)
            if m:
                data['入出口站'] = m.group(1).strip()
            else:
                m = re.search(r'入/出口站[：:]\s*(.+)', line)
                if m:
                    data['入出口站'] = m.group(1).strip()
            break

    # 开票人
    m = re.search(r'开票人[：:]\s*(\S+)', text)
    if m:
        data['开票人'] = m.group(1).strip()

    return data


def extract_field_safe(data, key, default=''):
    """安全提取字段"""
    return data.get(key, default)


# ========== PDF 处理 ==========


def process_pdf(pdf_path):
    """处理单个PDF文件，返回解析后的数据"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ''
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + '\n'
            if not text.strip():
                return {'error': '无法提取文本内容'}, pdf_path
            data = parse_invoice_text(text)
            data['文件名'] = Path(pdf_path).name
            return data, pdf_path
    except Exception as e:
        return {'error': str(e)}, pdf_path


def scan_pdf_files(folder_path):
    """扫描文件夹下所有PDF文件"""
    pdf_files = []
    for f in sorted(os.listdir(folder_path)):
        if f.lower().endswith('.pdf'):
            pdf_files.append(os.path.join(folder_path, f))
    return pdf_files


# ========== Excel 导出 ==========


def export_to_excel(results, output_path):
    """导出结果到Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '发票信息'

    # 定义表头
    headers = [
        '文件名',
        '发票号码',
        '开票日期',
        '购买方名称',
        '购买方纳税人识别号',
        '销售方名称',
        '销售方纳税人识别号',
        '项目名称',
        '数量',
        '单价',
        '金额',
        '价税合计',
        '价税合计大写',
        '车牌号',
        '车辆类型',
        '通行日期起止',
        '入出口站',
        '开票人',
        '文件路径',
    ]

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='微软雅黑', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, (data, pdf_path) in enumerate(results, 2):
        values = [
            extract_field_safe(data, '文件名'),
            extract_field_safe(data, '发票号码'),
            extract_field_safe(data, '开票日期'),
            extract_field_safe(data, '购买方名称'),
            extract_field_safe(data, '购买方纳税人识别号'),
            extract_field_safe(data, '销售方名称'),
            extract_field_safe(data, '销售方纳税人识别号'),
            extract_field_safe(data, '项目名称'),
            extract_field_safe(data, '数量'),
            extract_field_safe(data, '单价'),
            extract_field_safe(data, '金额'),
            extract_field_safe(data, '价税合计'),
            extract_field_safe(data, '价税合计大写'),
            extract_field_safe(data, '车牌号'),
            extract_field_safe(data, '车辆类型'),
            extract_field_safe(data, '通行日期起止'),
            extract_field_safe(data, '入出口站'),
            extract_field_safe(data, '开票人'),
            str(pdf_path),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 设置列宽
    col_widths = {
        1: 30,
        2: 24,
        3: 16,
        4: 30,
        5: 24,
        6: 30,
        7: 24,
        8: 18,
        9: 8,
        10: 10,
        11: 10,
        12: 10,
        13: 20,
        14: 12,
        15: 10,
        16: 28,
        17: 30,
        18: 12,
        19: 50,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动筛选
    ws.auto_filter.ref = f'A1:{chr(64 + len(headers))}1'

    wb.save(output_path)
    return output_path


# ========== GUI ==========


class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title('数电发票PDF识别导出工具')
        self.root.geometry('800x600')
        self.root.minsize(700, 500)

        # 设置样式
        self.root.option_add('*Font', '微软雅黑 10')

        self.selected_folder = tk.StringVar()
        self.status_text = tk.StringVar(value='就绪')

        self.results = []  # 存储所有解析结果
        self.build_ui()

        # 协议
        self.root.protocol('WM_DELETE_WINDOW', self.on_close)

    def build_ui(self):
        # 顶部：文件夹选择
        frame_top = ttk.Frame(self.root, padding=10)
        frame_top.pack(fill=tk.X)

        ttk.Label(frame_top, text='选择文件夹：').pack(side=tk.LEFT)

        self.folder_entry = ttk.Entry(frame_top, textvariable=self.selected_folder, width=60)
        self.folder_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(frame_top, text='浏览...', command=self.browse_folder).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_top, text='开始识别', command=self.start_scan).pack(side=tk.LEFT, padx=5)

        # 操作按钮区
        frame_actions = ttk.Frame(self.root, padding=(10, 0, 10, 5))
        frame_actions.pack(fill=tk.X)

        ttk.Label(frame_actions, text='扫描结果：').pack(side=tk.LEFT)

        self.btn_export = ttk.Button(
            frame_actions, text='导出到Excel', command=self.export_excel, state=tk.DISABLED
        )
        self.btn_export.pack(side=tk.RIGHT, padx=2)

        self.btn_clear = ttk.Button(
            frame_actions, text='清空结果', command=self.clear_results, state=tk.DISABLED
        )
        self.btn_clear.pack(side=tk.RIGHT, padx=2)

        # 进度条
        self.progress = ttk.Progressbar(self.root, mode='determinate')
        self.progress.pack(fill=tk.X, padx=10, pady=(0, 5))

        # 结果表格
        frame_table = ttk.Frame(self.root, padding=(10, 0, 10, 10))
        frame_table.pack(fill=tk.BOTH, expand=True)

        columns = ('文件名', '发票号码', '价税合计', '车牌号', '状态')
        self.tree = ttk.Treeview(frame_table, columns=columns, show='headings', height=15)

        self.tree.heading('文件名', text='文件名')
        self.tree.heading('发票号码', text='发票号码')
        self.tree.heading('价税合计', text='金额')
        self.tree.heading('车牌号', text='车牌号')
        self.tree.heading('状态', text='状态')

        self.tree.column('文件名', width=200)
        self.tree.column('发票号码', width=180)
        self.tree.column('价税合计', width=80)
        self.tree.column('车牌号', width=100)
        self.tree.column('状态', width=80)

        scrollbar = ttk.Scrollbar(frame_table, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 底部状态栏
        frame_status = ttk.Frame(self.root, padding=(10, 0, 10, 5))
        frame_status.pack(fill=tk.X)
        ttk.Label(frame_status, textvariable=self.status_text).pack(side=tk.LEFT)

        # 绑定双击事件查看详情
        self.tree.bind('<Double-1>', self.on_item_double_click)

    def browse_folder(self):
        folder = filedialog.askdirectory(title='选择包含PDF发票的文件夹')
        if folder:
            self.selected_folder.set(folder)

    def start_scan(self):
        folder = self.selected_folder.get().strip()
        if not folder:
            messagebox.showwarning('提示', '请先选择文件夹')
            return

        if not os.path.isdir(folder):
            messagebox.showerror('错误', '文件夹路径无效')
            return

        pdf_files = scan_pdf_files(folder)
        if not pdf_files:
            messagebox.showinfo('提示', '该文件夹下没有找到PDF文件')
            return

        # 清空上次结果
        self.clear_results()

        total = len(pdf_files)
        self.progress['maximum'] = total
        self.status_text.set(f'正在识别 {total} 个PDF文件...')
        self.root.update_idletasks()

        success_count = 0
        error_count = 0

        for i, pdf_path in enumerate(pdf_files):
            data, path = process_pdf(pdf_path)

            filename = Path(pdf_path).name

            if 'error' in data:
                status = '识别失败'
                error_count += 1
                self.tree.insert('', tk.END, values=(filename, '', '', '', status))
            else:
                status = '成功'
                success_count += 1
                self.results.append((data, pdf_path))
                self.tree.insert(
                    '',
                    tk.END,
                    values=(
                        filename,
                        data.get('发票号码', ''),
                        data.get('价税合计', ''),
                        data.get('车牌号', ''),
                        status,
                    ),
                )

            self.progress['value'] = i + 1
            self.status_text.set(f'正在识别 ({i + 1}/{total}): {filename}')
            self.root.update_idletasks()

        # 更新状态
        self.status_text.set(
            f'扫描完成：共 {total} 个文件，成功 {success_count} 个，失败 {error_count} 个'
        )

        if self.results:
            self.btn_export.config(state=tk.NORMAL)
        self.btn_clear.config(state=tk.NORMAL)

        if error_count > 0 and success_count == 0:
            messagebox.showerror('错误', '所有文件识别失败，请检查PDF是否为有效的数电发票格式')

    def export_excel(self):
        if not self.results:
            messagebox.showwarning('提示', '没有可导出的数据')
            return

        default_name = f'发票信息_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        folder = self.selected_folder.get().strip() or os.path.expanduser('~')
        output_path = filedialog.asksaveasfilename(
            title='保存Excel文件',
            initialdir=folder,
            initialfile=default_name,
            defaultextension='.xlsx',
            filetypes=[('Excel文件', '*.xlsx')],
        )

        if not output_path:
            return

        try:
            export_to_excel(self.results, output_path)
            messagebox.showinfo('成功', f'已导出 {len(self.results)} 条记录到：\n{output_path}')
            self.status_text.set(f'已导出到：{output_path}')
        except Exception as e:
            messagebox.showerror('导出失败', str(e))

    def clear_results(self):
        self.results.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.btn_export.config(state=tk.DISABLED)
        self.btn_clear.config(state=tk.DISABLED)
        self.progress['value'] = 0
        self.status_text.set('就绪')

    def on_item_double_click(self, event):
        selected = self.tree.selection()
        if not selected:
            return
        item = self.tree.item(selected[0])
        values = item['values']
        if len(values) >= 2 and values[1]:
            # 查找对应的数据
            filename = values[0]
            for data, pdf_path in self.results:
                if data.get('文件名') == filename:
                    detail = '\n'.join(f'{k}: {v}' for k, v in data.items() if v)
                    messagebox.showinfo(f'发票详情 - {filename}', detail)
                    return

    def on_close(self):
        self.root.destroy()


def main():
    root = tk.Tk()
    InvoiceApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
